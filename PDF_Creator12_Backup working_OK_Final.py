from PyQt5 import QtCore, QtGui, QtWidgets
import sys
import openpyxl

# from GUI2 import Ui_MainWindow
import os.path
import re
import threading
from PyQt5 import uic
import datetime,time,schedule



Ui_MainWindow, QtBaseClass = uic.loadUiType(r'PDF_Creator_GUI3.ui')

app = QtWidgets.QApplication(sys.argv)
app.setWindowIcon(QtGui.QIcon("icon.ico"))

MainWindow = QtWidgets.QMainWindow()

ui = Ui_MainWindow()
ui.setupUi(MainWindow)

def parse_table(table):

    #print("tabel show in parse funtion",table)
    df = table.replace(float("NaN"), "")
    if (df.shape[1] < 3):
        print("Enter in DF shape if")
        return (pd.DataFrame())
    print("Dataframe shape",df.shape)
    print("ataframe df type is ",type(df))
    # print("DF print",df)

    # print('DF..... Optional circuit no column',df.loc[:,"Optional Circuit No."])
    # print('DF..... Page no column',df.loc[:,"▲3Page No."])
    # print("Column value in text",list(df.columns.values))
    list1=list(df.columns.values)
    print("number of columns in table",list1)
    # number_columns_len=len(list(df.columns.values))

    df4 = pd.Series(list1)
    # j = 0

    # while j < number_columns_len:
    #     print(list1[j])
       
    #     j += 1
    # for col in df.columns:
    #     print("getting only dataframe column name",col)
    #     print("type of col",type(col))
    #     if (col.str.contains(r'OPTIONAL CIRCUIT NO.',case=False)).any():
    #         circuit_no_col=col
    #         print("circuit_no_col through column ",circuit_no_col)
    
    
    try: # find table with group7 in page

        print("Enter in parsed try ")
        if(df4[df4.str.contains(r'OPTIONAL CIRCUIT NO',case=False)].any()):
            print("enter in optional circuit&&&&&&&&&")
            circuit_no_col =df4[df4.str.contains(r'OPTIONAL CIRCUIT NO',case=False)].values[0]
            print("Final circuit name is",circuit_no_col)

        if(df4[df4.str.contains(r'GROUP',case=False)].any()):
            group_col =df4[df4.str.contains(r'GROUP',case=False)].values[0]
            print("Final  group name is",group_col)
        if(ui.checkBox_2_M_Series.isChecked() == True):

            if(df4[df4.str.contains(r'SHEET',case=False)].any()):
                page_col =df4[df4.str.contains(r'SHEET',case=False)].values[0]
                print("Final  page name is",page_col) 
        if(ui.checkBox_2_M_Series.isChecked() == False):  
             if(df4[df4.str.contains(r'Page No.',case=False)].any()):
                page_col =df4[df4.str.contains(r'Page No.',case=False)].values[0]
                print("Final  page name is",page_col) 


        # if(df4[df4.str.contains(r'▲',case=False)].any()):
        #     # page_col =df4[df4.str.contains(r'SHEET',case=False)].values[0]
        #     page_col="▲"
        #     print("Final  page name is",page_col)         


        # for i in df:
        #     print("df[i]@@@@@@@@@@@@@@@@@@@@",i)
        #         # print("Type of df[i] is ~~~~~~~",type(df[i]))
        #     if (df[i].str.contains(r'OPTIONAL CIRCUIT NO.',case=False)).any():
        #         circuit_no_col = i
        #         # circuit_no_col = "Optional Circuit No."
        #         print("circuit_no_col is.....################",circuit_no_col)
        #     if (df[i].str.contains(r'GROUP')).any():
        #         group_col = i
        #         # group_col = "Group"
        #         print("group_col is.....####################",group_col)
        #     if (df[i].str.contains(r'▲3 Page No.')).any():
        #         page_col = i
        #         # page_col ="▲3Page No."
        #         print("page_col is.....##########",page_col,i)

        if circuit_no_col and page_col and group_col:
            print("Enter in ~~~~~~~if statement")
            print("Dataframe with new column name ~~~~~~~~~~~~~~~~~~~~~~--------------",df[[group_col, circuit_no_col, page_col]])
            new_dataframe = df[[group_col, circuit_no_col, page_col]].replace("", float("NaN")).dropna().reset_index(drop = True)
            
            print("New dataframe **************",new_dataframe)
            new_dataframe.columns = ['Group', 'Circuit no', 'Page no']
            print("New dataframe after column change###########",new_dataframe.columns)  

            if(ui.checkBox_2_M_Series.isChecked() == True):       
                option_pages = new_dataframe[new_dataframe["Group"].isin(["GROUP 9 - MOLD","GROUP 10 -INTERFACE","GROUP 11 - CUSTOM OPTIONS"])]           
            if(ui.checkBox_2_M_Series.isChecked() == False): 
                option_pages = new_dataframe[new_dataframe["Group"] == "GROUP 7"]
            print("option pages%%%%%%%%%%%%%%%%%%%",option_pages)
            option_pages.reset_index(drop=True)
            
            option_pages['Page no']=option_pages['Page no'].astype(str)
            
           
            
            parsed_df = pd.DataFrame(data = np.array([list(option_pages["Group"]), list(option_pages['Circuit no']), list(option_pages["Page no"].apply(lambda x: ((re.findall(r'\d+', x)[0]))))]).T, columns = ['Group', 'Circuit no', 'Page no'])
            print("parsedf``````````````````",parsed_df)
            return (parsed_df)
        else:
            return (pd.DataFrame())
    
    except Exception as e:    
        print("Enter in perse except.....>>>>>>>>>",e)
        return (pd.DataFrame())



def delete_pages():

    global stop

    stop = False
    #doc = fitz.open("11831761_D.pdf")
    #del_pag_list = []
    # try:

    print("enter in delete 1")
    import pandas as pd
    # toc_file_name = ui.lineEdit.text()
    try:
        file_name = ui.lineEdit_2.text()
        print("read file name",file_name)
        if (os.path.exists(file_name)):

    #---------------------------------------------------------Get Circuit number----------------------------
           

            spl_word = '/'
            Get_last_word = file_name.split(spl_word)[-1]

            Circuit_NO=Get_last_word[0:8]
            print("Circuit number:",Circuit_NO)


            ui.textEdit_2.append(f"Circuit_NO:{Circuit_NO} PDF generation Start........") 

            # Find pages that have index
            object = PyPDF2.PdfFileReader(file_name)
            # print("Object for test",object)
            if(ui.checkBox_2_M_Series.isChecked() == True): 
            # String = "/ Index"
                String = "TABLE OF CONTENTS"
            if(ui.checkBox_2_M_Series.isChecked() == False): 
                String = "/ Index"
            

            if(ui.checkBox_2_M_Series.isChecked() == True): 
                page_read=8
            if(ui.checkBox_2_M_Series.isChecked() == False): 
                page_read=12

            index_pages = []
            for i in range(0, page_read):
                PageObj = object.getPage(i)
                Text = PageObj.extractText()
                # print("file text @@@@@@@@@@@@$$$$$$$!!!!!!!!~~~~~",Text)
                ResSearch = re.search(String, Text)

                print("Research is.......",ResSearch)
                if ResSearch:
                    index_pages.append(i+1)
                   
                    print("Index page",index_pages)
            # index_pages = str(index_pages).replace('[', "").replace(']', "")
            object.stream.close()
            # Get Page size

            input1 = PyPDF2.PdfFileReader(open(file_name, 'rb'))

            size = input1.getPage(index_pages[0]).mediaBox

            print("Size of page 0",size)
            input1.stream.close()
            # read tables from index pages
            # margin = 30
            # tables = tabula.read_pdf(file_name, pages=index_pages, multiple_tables=True , lattice = True, guess = False, area = [size[0]+20, size[1]+30, size[3]-60, size[2]-50])
            tables = tabula.read_pdf(file_name, pages=index_pages, multiple_tables=True,stream=True)# java is found but command is not working, try java 7

            option_page_dataframe = pd.DataFrame(columns = ['Group', 'Circuit no', 'Page no'])

            # print("opton page dataframe",option_page_dataframe)

            for table in tables:
                
                parsed = parse_table(table)

                
                # print("parsed data after funtion return",parsed)
                if not (parsed.empty):
                    print("enter in parsed section")
                    option_page_dataframe = pd.concat([option_page_dataframe, parsed])
                    
            option_page_dataframe.reset_index(inplace=True)
            #option_page_dataframe.to_csv('file1.csv')
            print("option_page_dataframe after ####$%#%%%%%%....",option_page_dataframe)

            # keep_options = int(ui.textEdit.toPlainText().split())
            # inputSonumber = list(map(int, ui.textEdit_Input.toPlainText().split()))
            keep_options = list(map(int, ui.textEdit.toPlainText().split()))
            print("Keep options",keep_options)
            delete_page_dataframe = option_page_dataframe.copy()
            



            if len(keep_options) != 0:
                delete_page_dataframe.dropna(axis=0, inplace=True)

            for option in keep_options:


                if ((pd.to_numeric(delete_page_dataframe["Circuit no"]) == int(option)).any()):
                    print("Enter in delete loop&&&&&&&&&&))))))")
                    print(delete_page_dataframe["Circuit no"])
                    print("Only circut no column",delete_page_dataframe["Circuit no"])
                    print("Deleter page dataframe before^^^^^^^^",delete_page_dataframe)
                    delete_page_dataframe[['Circuit no', 'Page no']] = delete_page_dataframe[['Circuit no', 'Page no']].apply(pd.to_numeric)
                    
                    delete_page_dataframe.drop(delete_page_dataframe[delete_page_dataframe["Circuit no"] == int(option)].index, inplace=True)
                    print("Deleter page dataframe after-----------------",delete_page_dataframe)
                else:

                    ui.textEdit_2.setTextColor(redColor)
                    ui.textEdit_2.append(str(option) + "  not found in index.")
                    ui.textEdit_2.setTextColor(greenColor)

                # delete_page_dataframe.drop(delete_page_dataframe[delete_page_dataframe['Circuit no'] == option].index, inplace=True) #Old code hide

            # del_pag_list = list(delete_page_dataframe["Page no"]) #Old code hide
            delete_page_dataframe[['Circuit no', 'Page no']] = delete_page_dataframe[['Circuit no', 'Page no']].apply(pd.to_numeric)
            # del_pag_list = sorted(list(map(int, del_pag_list)), reverse=True)#Old code hide
            print("for loop outer delete page dtafrmae !!!!!!!!!",delete_page_dataframe)
            del_pag_list=delete_page_dataframe["Page no"].tolist() #New line adde change list funtion
            print("Delete page list before sort #########*********  ",del_pag_list)
            del_pag_list=sorted(del_pag_list,reverse=True) # New line added for sort fntion this solve bad page arrer with options

            print("Delete page list After sort (((((*********))))))) ",del_pag_list)
            del_pag_list = [int(item) for item in del_pag_list]#New line added for convert list to integer
            print("Delete page list @@@@@@@@@@@@@  ",del_pag_list)
            ui.textEdit_2.append("Deleted Pages - " + str(len(del_pag_list)) + "\n" + str(del_pag_list))



#---------------------------------------------------------------History Sheet save------------------------------------------
            x1=datetime.datetime.now()
            Time=x1.strftime("%d-%m-%y %H:%M:%S")

            Ckt_Page_Delete_history_file=openpyxl.load_workbook("Ckt_Page_Delete_History.xlsx")
            sheet=Ckt_Page_Delete_history_file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=Circuit_NO) # Circuit Number
            sheet.cell(column=2,row=sheet.max_row,value=str(len(del_pag_list))) #No Of page delete
            sheet.cell(column=3,row=sheet.max_row,value=Time) #Time
            Ckt_Page_Delete_history_file.save("Ckt_Page_Delete_History.xlsx")
#-------------------------------------------------------------------------------------------------------------------------------------------
            doc = fitz.open(file_name)
            # for page in del_pag_list:
            #     doc.delete_page(int(page) - 1)
            # doc.delete_pages(del_pag_list)
            # numbers = (1, 2, 3, 4)
            # result = map(lambda x: x-1, del_pag_list)
            print("One page reduce list",(list(map(lambda x: x-1, del_pag_list))))
            doc.delete_pages(list(map(lambda x: x-1, del_pag_list)))
           
            if(stop==True):
                ui.pushButton.setEnabled(True)
                return
            if(ui.checkBox.isChecked() == False):
                try:

                    doc.save(os.path.dirname(file_name) + "/temp.pdf")
                    doc.close()
                    os.remove(file_name)
                    os.rename(os.path.dirname(file_name) + "/temp.pdf" , file_name)
                    os.startfile(file_name)
                except:
                    ui.textEdit_2.setTextColor(blueColor)
                    ui.textEdit_2.append(os.path.basename(file_name) + " File open in other service, please close it first.")
                    ui.textEdit_2.setTextColor(greenColor)
            else:
                try:

                    save_path = ui.lineEdit_5.text()
                    print("Output path",save_path)
                    if not(save_path == "" or os.path.exists(save_path)==False):

                        base_name = os.path.basename(file_name)
                        save_path = ui.lineEdit_5.text() + r"\\" + base_name
                        doc.save(os.path.dirname(save_path) + "/temp.pdf")
                        doc.close()
                        #os.remove(file_name)
                        os.rename(os.path.dirname(save_path) + "/temp.pdf", save_path)
                        os.startfile(save_path)

                    else:
                        ui.textEdit_2.setTextColor(blueColor)
                        ui.textEdit_2.append(os.path.basename("Enter valid save path"))
                        ui.textEdit_2.setTextColor(greenColor)
                except:
                    ui.textEdit_2.setTextColor(blueColor)
                    ui.textEdit_2.append(os.path.basename(file_name) + " File open in other service, please close it first.")
                    ui.textEdit_2.setTextColor(greenColor)

        else:

            ui.textEdit_2.setTextColor(blueColor)
            ui.textEdit_2.append("PDF file dose not exist")
            ui.textEdit_2.setTextColor(greenColor)


    except Exception as e:

        ui.textEdit_2.setTextColor(redColor)
        ui.textEdit_2.append(str(e))
        ui.textEdit_2.append(str("Enter in exeptions"))
        ui.textEdit_2.setTextColor(greenColor)

    ui.pushButton.setEnabled(True)
    return


def clear_test():
    ui.textEdit.clear()
    ui.textEdit_2.clear()
    # global stop
    # stop = True
    
    browse2()
    

'''
for i in range(60-min(60, len(ch_7))):
    ui.emptyLabels.append(QtWidgets.QLabel())
    ui.verticalLayout_2.addWidget(ui.emptyLabels[i])
'''
# def Message_display():
#     print("PDF generation start...@@@@@@@@@.......")
#     ui.textEdit_2.append("PDF generation start.............")
#     ui.textEdit_2.setTextColor(greenColor)
    


def browse3():
    path = QtWidgets.QFileDialog.getExistingDirectory(directory="")
    ui.lineEdit_5.setText(path)


def browse2():
    path = QtWidgets.QFileDialog.getOpenFileName(filter='*.pdf')[0]
    ui.lineEdit_2.setText(path)

def set_logo():
    label = QtWidgets.QLabel(ui.frame)
    pixmap = QtGui.QPixmap("Milacron_logo.png")
    pixmap_resized = pixmap.scaled(180, 200, QtCore.Qt.KeepAspectRatio)
    label.setPixmap(pixmap_resized)
    label.show()

def enable_save_to_different_path():
    if ui.checkBox.isChecked():
        ui.lineEdit_5.setEnabled(True)
        ui.pushButton_8.setEnabled(True)

    else:
        ui.lineEdit_5.setEnabled(False)
        ui.pushButton_8.setEnabled(False)

def create_pdf_clicked():
    ui.pushButton.setEnabled(False)
    # message_display_thread=threading.Thread(target=Message_display)
    # message_display_thread.start()
    # message_display_thread.join()
    delete_pdf_thread = threading.Thread(target=delete_pages) # define thread
    delete_pdf_thread.start()
    print("Enter in pdf clicked funtion ..................................................................................")

    # delete_pdf_thread.join()

def stop_the_process():

    global stop
    stop = True

    ui.textEdit_2.setTextColor(blueColor)
    ui.textEdit_2.append("PDF generation stopped.")
    ui.textEdit_2.setTextColor(greenColor)

import fitz
import tabula
import PyPDF2
import pandas as pd
import numpy as np
sys.setrecursionlimit(10000)

# file_name  = '11831761_F.pdf'

# ui.pushButton.clicked.connect(delete_pages)
ui.pushButton.clicked.connect(create_pdf_clicked)

ui.pushButton_3.clicked.connect(clear_test)
# ui.pushButton_2.clicked.connect(browse)
ui.pushButton_4.clicked.connect(browse2)
ui.pushButton_8.clicked.connect(browse3)

ui.pushButton_5.clicked.connect(stop_the_process)
ui.pushButton_5.setEnabled(False)


ui.checkBox.clicked.connect(enable_save_to_different_path)

####################################################################################################
ui.textEdit_2.setReadOnly(True)

redColor = QtGui.QColor(255, 0, 0)
greenColor = QtGui.QColor(70,170,50)
blueColor = QtGui.QColor(0,0,255)

ui.textEdit_2.setTextColor(greenColor)

set_logo()

ui.lineEdit_5.setEnabled(False)
ui.pushButton_8.setEnabled(False)

# ui.label.setText("Index file")


cursor = ui.textEdit.textCursor()
cursor.movePosition(cursor.Left, cursor.KeepAnchor,  3)

# delete_pdf_thread = threading.Thread(target=delete_pages) # define thread
# delete_pdf_thread.daemon = True


stop = False

MainWindow.setFixedSize(MainWindow.size())

MainWindow.show()
sys.exit(app.exec_())

#pyinstaller --icon=icon.ico --add-data "Milacron_logo.ico;." --add-data "Settings_logo.png;." --noconsole --name PDF_Comparator GUI2.py

#pyinstaller --icon=icon.ico --add-data "icon.ico;." --noconsole --name PDF_Creator PDF_Creator4.py

#pyinstaller --icon=icon.ico --add-data "icon.ico;." --noconsole --add-data "Milacron_logo.png;." --add-data "tabula" --name PDF_Creator PDF_Creator6.py
